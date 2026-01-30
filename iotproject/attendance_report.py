from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

def generate_excel_report(attendance_manager, student_db):
    """Generate comprehensive Excel attendance report"""
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"attendance_report_{timestamp}.xlsx"
    
    wb = Workbook()
    
    # ----------------------------------------------------------------------
    # SHEET 1: TODAY'S ATTENDANCE
    # ----------------------------------------------------------------------
    ws_today = wb.active
    ws_today.title = "Today's Attendance"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    # Report title
    ws_today['A1'] = "DAILY ATTENDANCE REPORT"
    ws_today['A1'].font = Font(bold=True, size=16)
    ws_today['A2'] = f"Date: {datetime.now().strftime('%Y-%m-%d')}"
    ws_today['A2'].font = Font(size=11)
    
    # Column headers
    headers = ["Sr. No.", "Roll Number", "Name", "Time", "Status"]
    for col, header in enumerate(headers, start=1):
        cell = ws_today.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Get today's attendance
    today_attendance = attendance_manager.get_today_attendance()
    
    # Data rows
    for idx, record in enumerate(today_attendance, start=1):
        row = idx + 4
        ws_today.cell(row=row, column=1, value=idx)
        ws_today.cell(row=row, column=2, value=record['roll_number'])
        ws_today.cell(row=row, column=3, value=record['name'])
        ws_today.cell(row=row, column=4, value=record['timestamp'].split()[1])  # Time only
        ws_today.cell(row=row, column=5, value="Present")
        
        # Status cell formatting
        status_cell = ws_today.cell(row=row, column=5)
        status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        status_cell.font = Font(color="006100")
    
    # Summary section
    summary_row = len(today_attendance) + 6
    ws_today.cell(row=summary_row, column=1, value="SUMMARY")
    ws_today.cell(row=summary_row, column=1).font = Font(bold=True, size=12)
    
    ws_today.cell(row=summary_row + 1, column=1, value="Total Students:")
    ws_today.cell(row=summary_row + 1, column=2, value=len(student_db.students))
    
    ws_today.cell(row=summary_row + 2, column=1, value="Present:")
    ws_today.cell(row=summary_row + 2, column=2, value=len(today_attendance))
    ws_today.cell(row=summary_row + 2, column=2).font = Font(color="006100", bold=True)
    
    ws_today.cell(row=summary_row + 3, column=1, value="Absent:")
    absent_count = len(student_db.students) - len(today_attendance)
    ws_today.cell(row=summary_row + 3, column=2, value=absent_count)
    ws_today.cell(row=summary_row + 3, column=2).font = Font(color="9C0006", bold=True)
    
    ws_today.cell(row=summary_row + 4, column=1, value="Attendance %:")
    if len(student_db.students) > 0:
        attendance_pct = f"=B{summary_row + 2}/B{summary_row + 1}"
        ws_today.cell(row=summary_row + 4, column=2, value=attendance_pct)
        ws_today.cell(row=summary_row + 4, column=2).number_format = '0.0%'
    
    # Column widths
    ws_today.column_dimensions['A'].width = 10
    ws_today.column_dimensions['B'].width = 15
    ws_today.column_dimensions['C'].width = 25
    ws_today.column_dimensions['D'].width = 15
    ws_today.column_dimensions['E'].width = 12
    
    # ----------------------------------------------------------------------
    # SHEET 2: FULL ATTENDANCE LOG
    # ----------------------------------------------------------------------
    ws_log = wb.create_sheet("Full Attendance Log")
    
    # Headers
    headers_log = ["Sr. No.", "Roll Number", "Name", "Date", "Time", "Day"]
    for col, header in enumerate(headers_log, start=1):
        cell = ws_log.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # All attendance records
    all_records = attendance_manager.attendance_log
    for idx, record in enumerate(all_records, start=1):
        row = idx + 1
        date_obj = datetime.strptime(record['date'], '%Y-%m-%d')
        day_name = date_obj.strftime('%A')
        
        ws_log.cell(row=row, column=1, value=idx)
        ws_log.cell(row=row, column=2, value=record['roll_number'])
        ws_log.cell(row=row, column=3, value=record['name'])
        ws_log.cell(row=row, column=4, value=record['date'])
        ws_log.cell(row=row, column=5, value=record['timestamp'].split()[1])
        ws_log.cell(row=row, column=6, value=day_name)
    
    # Column widths
    ws_log.column_dimensions['A'].width = 10
    ws_log.column_dimensions['B'].width = 15
    ws_log.column_dimensions['C'].width = 25
    ws_log.column_dimensions['D'].width = 15
    ws_log.column_dimensions['E'].width = 12
    ws_log.column_dimensions['F'].width = 12
    
    # ----------------------------------------------------------------------
    # SHEET 3: STUDENT MASTER LIST
    # ----------------------------------------------------------------------
    ws_students = wb.create_sheet("Student List")
    
    # Headers
    headers_students = ["Sr. No.", "Roll Number", "Name", "Registration Date"]
    for col, header in enumerate(headers_students, start=1):
        cell = ws_students.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Student data
    students_list = sorted(student_db.students.items(), key=lambda x: x[0])
    for idx, (roll, data) in enumerate(students_list, start=1):
        row = idx + 1
        ws_students.cell(row=row, column=1, value=idx)
        ws_students.cell(row=row, column=2, value=roll)
        ws_students.cell(row=row, column=3, value=data['name'])
        ws_students.cell(row=row, column=4, value=data.get('added_date', 'N/A'))
    
    # Column widths
    ws_students.column_dimensions['A'].width = 10
    ws_students.column_dimensions['B'].width = 15
    ws_students.column_dimensions['C'].width = 25
    ws_students.column_dimensions['D'].width = 20
    
    # ----------------------------------------------------------------------
    # SAVE WORKBOOK
    # ----------------------------------------------------------------------
    wb.save(filename)
    print(f"\n✅ Excel Report Generated: {filename}")
    print(f"   • Today's Attendance: {len(today_attendance)} students")
    print(f"   • Total Records: {len(all_records)}")
    print(f"   • Registered Students: {len(student_db.students)}")
    
    return filename

if __name__ == "__main__":
    # Test report generation
    from classroom_attendance import StudentDatabase, AttendanceManager
    
    print("Generating test report...")
    student_db = StudentDatabase()
    attendance_manager = AttendanceManager()
    
    if len(student_db.students) > 0:
        generate_excel_report(attendance_manager, student_db)
    else:
        print("⚠️  No students in database. Please run setup_students.py first.")